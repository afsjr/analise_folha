import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

# Carregar dados completos
with open('data/folha_completa.json', 'r', encoding='utf-8') as f:
    full_data = json.load(f)

# Agrupar dados por mês
meses = {}
for d in full_data:
    mes = d.get('mes', '2026-04')
    if mes not in meses:
        meses[mes] = []
    meses[mes].append(d)

header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Aba Evolução Mensal
ws = wb.create_sheet(title='Evolução')
ws['A1'] = 'EVOLUÇÃO MENSAL DA FOLHA'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

ws['A3'] = 'Mês'
ws['B3'] = 'Funcionários'
ws['C3'] = 'Total Bruto'
ws['D3'] = 'Total Líquido'
ws['E3'] = 'Variação'
for c in range(1, 6):
    cell = ws.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

mes_keys = sorted(meses.keys())
r = 4
anterior = 0
for mes in mes_keys:
    dados = meses[mes]
    brute = sum(d['total_bruto'] for d in dados)
    liqui = sum(d['total_liquido'] for d in dados)
    varicao = ((brute - anterior) / anterior * 100) if anterior > 0 else 0
    
    ws.cell(row=r, column=1, value=mes)
    ws.cell(row=r, column=2, value=len(dados))
    ws.cell(row=r, column=3, value=brute)
    ws.cell(row=r, column=4, value=liqui)
    ws.cell(row=r, column=5, value=varicao / 100)
    
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = thin_border
    
    anterior = brute
    r += 1

ws['A' + str(r)] = 'TOTAL'
ws['A' + str(r)].font = Font(bold=True)
ws['C' + str(r)] = f'=SUM(C4:C{r-1})'
ws['D' + str(r)] = f'=SUM(D4:D{r-1})'
for c in range(1, 6):
    ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=c).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12

# Abas por segmento (mês atual)
with open('data/folha.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

segmentos = {}
for item in data:
    seg = item.get('segmento', 'Outros')
    if seg not in segmentos:
        segmentos[seg] = []
    segmentos[seg].append(item)

for seg, itens in sorted(segmentos.items()):
    nome_aba = seg[:31]
    ws = wb.create_sheet(title=nome_aba)
    
    cols = ['Nome', 'Cargo', 'Salário', 'Adic. Grad', 'Adic. Pós', 'Sal. Família', 'Gratif 5%', 'Total Bruto', 'INSS', 'Créditos', 'Total Líquido']
    
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for row_idx, item in enumerate(itens, 2):
        ws.cell(row=row_idx, column=1, value=item['nome'])
        ws.cell(row=row_idx, column=2, value=item['cargo'])
        ws.cell(row=row_idx, column=3, value=item['salario'])
        ws.cell(row=row_idx, column=4, value=item['adicional_graduacao'])
        ws.cell(row=row_idx, column=5, value=item['adicional_pos'])
        ws.cell(row=row_idx, column=6, value=item['salario_familia'])
        ws.cell(row=row_idx, column=7, value=item['gratificacao_5'])
        ws.cell(row=row_idx, column=8, value=item['total_bruto'])
        ws.cell(row=row_idx, column=9, value=item['inss'])
        ws.cell(row=row_idx, column=10, value=item['creditos_extras'])
        ws.cell(row=row_idx, column=11, value=item['total_liquido'])
        
        for c in range(1, 12):
            ws.cell(row=row_idx, column=c).border = thin_border
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    for col in range(3, 12):
        ws.column_dimensions[chr(64 + col)].width = 11

# Aba Resumo
ws = wb.create_sheet(title='Resumo')
ws['A1'] = 'RESUMO POR SEGMENTO - ABRIL 2026'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

ws['A3'] = 'Segmento'
ws['B3'] = 'Funcionários'
ws['C3'] = 'Total Bruto'
ws['D3'] = 'Total Líquido'
ws['E3'] = '% Folha'
for c in range(1, 6):
    cell = ws.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

r = 4
total_bruto = sum(s['total_bruto'] for s in data)
for seg in sorted(segmentos.keys()):
    itens = segmentos[seg]
    brute = sum(i['total_bruto'] for i in itens)
    liqui = sum(i['total_liquido'] for i in itens)
    ws.cell(row=r, column=1, value=seg)
    ws.cell(row=r, column=2, value=len(itens))
    ws.cell(row=r, column=3, value=brute)
    ws.cell(row=r, column=4, value=liqui)
    ws.cell(row=r, column=5, value=brute/total_bruto)
    r += 1

ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
ws.cell(row=r, column=2, value=len(data))
ws.cell(row=r, column=3, value=total_bruto)
ws.cell(row=r, column=4, value=sum(s['total_liquido'] for s in data))
ws.cell(row=r, column=5, value=1)

for c in range(1, 6):
    ws.cell(row=r, column=c).border = thin_border
    ws.cell(row=r, column=c).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 10

wb.save('resultados/folha_consolidada.xlsx')
print(f"Excel salvo: resultados/folha_consolidada.xlsx")